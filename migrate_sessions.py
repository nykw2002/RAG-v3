#!/usr/bin/env python3
"""
Migrate existing session data from JSON files to Excel tracking spreadsheet
Avoids duplicates and handles various data formats
"""

import json
import os
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import hashlib

def get_system_prompt_hash(system_prompt):
    """Create a short hash of the system prompt for tracking"""
    if not system_prompt:
        return "N/A"
    # Create hash and add first few words for identification
    hash_val = hashlib.md5(system_prompt.encode()).hexdigest()[:8]
    # Get first meaningful words (skip common starting words)
    words = system_prompt.split()
    meaningful_words = []
    skip_words = {'you', 'are', 'an', 'ai', 'assistant', 'the', 'a', 'and', 'or', 'but', 'in', 'on', 'at', 'to', 'for', 'of', 'with', 'by'}
    
    for word in words[:20]:  # Check first 20 words
        clean_word = word.lower().strip('.,!?;:"')
        if len(clean_word) > 3 and clean_word not in skip_words:
            meaningful_words.append(clean_word)
        if len(meaningful_words) >= 3:
            break
    
    if meaningful_words:
        return f"{hash_val} ({', '.join(meaningful_words)})"
    else:
        return f"{hash_val} (prompt)"

def get_system_prompt_version(system_prompt):
    """Get a human-readable version identifier for the system prompt"""
    if not system_prompt:
        return "N/A"
    
    prompt_lower = system_prompt.lower()
    
    # Identify different prompt versions based on key phrases
    if "two-phase process" in prompt_lower:
        return "v2.0 (Two-Phase)"
    elif "critical rule" in prompt_lower:
        return "v1.5 (Critical Rule)"
    elif "script results are authoritative" in prompt_lower:
        return "v1.2 (Authoritative)"
    elif "query_complete" in prompt_lower:
        return "v1.1 (Query Complete)"
    elif "specialized in analyzing" in prompt_lower:
        return "v1.0 (Basic)"
    else:
        # Fallback to hash for unknown versions
        return get_system_prompt_hash(system_prompt)

def extract_script_content(iterations):
    """Extract the actual script content from iterations"""
    all_scripts = []
    
    for i, iteration in enumerate(iterations, 1):
        if iteration.get('script_executed', False):
            exec_result = iteration.get('execution_result', {})
            script_content = exec_result.get('script_content', '')
            if script_content:
                script_header = f"\n=== Iteration {i} Script ==="
                all_scripts.append(script_header)
                all_scripts.append(script_content)
    
    if all_scripts:
        return '\n'.join(all_scripts)
    else:
        return "No scripts executed"

def clean_text_for_excel(text):
    """Clean text to be Excel-friendly"""
    if not text:
        return ""
    
    # Convert to string and limit length for Excel readability
    text = str(text).strip()
    
    # Replace problematic characters
    text = text.replace('\n', ' | ').replace('\r', '')
    
    # Limit length to avoid Excel issues
    if len(text) > 1000:
        text = text[:997] + "..."
    
    return text

def migrate_sessions_to_excel():
    """Main migration function"""
    
    # Paths
    sessions_dir = Path("query_sessions")
    excel_path = Path("query_tracking.xlsx")
    
    if not sessions_dir.exists():
        print("No query_sessions directory found")
        return
    
    if not excel_path.exists():
        print("Excel tracking file not found. Please create it first.")
        return
    
    # Load existing Excel file
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        print(f"Loaded Excel file: {excel_path}")
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return
    
    # Get existing session IDs to avoid duplicates
    existing_sessions = set()
    for row in range(2, ws.max_row + 1):
        session_id = ws.cell(row=row, column=1).value
        if session_id:
            existing_sessions.add(session_id)
    
    print(f"Found {len(existing_sessions)} existing entries in Excel")
    
    # Find all JSON session files
    json_files = list(sessions_dir.glob("session_*.json"))
    print(f"Found {len(json_files)} JSON session files")
    
    migrated_count = 0
    skipped_count = 0
    error_count = 0
    
    for json_file in json_files:
        try:
            # Load session data
            with open(json_file, 'r', encoding='utf-8') as f:
                session_data = json.load(f)
            
            session_id = session_data.get('session_id', json_file.stem)
            
            # Skip if already exists
            if session_id in existing_sessions:
                skipped_count += 1
                continue
            
            # Extract data for Excel
            user_prompt = clean_text_for_excel(session_data.get('user_query', ''))
            
            # System prompt handling - now stored in JSON  
            system_prompt = session_data.get('system_prompt', '')
            system_prompt_full = clean_text_for_excel(system_prompt)  # Full text
            
            # Files accessed
            files_accessed = session_data.get('files_accessed', [])
            if isinstance(files_accessed, list):
                files_str = ", ".join(files_accessed) if files_accessed else "None detected"
            else:
                files_str = str(files_accessed)
            files_str = clean_text_for_excel(files_str)
            
            # Number of iterations
            num_iterations = session_data.get('total_iterations', 0)
            
            # Script content - actual code
            iterations = session_data.get('iterations', [])
            script_content = extract_script_content(iterations)
            script_content_clean = clean_text_for_excel(script_content)
            
            # AI final output
            final_output = clean_text_for_excel(session_data.get('final_answer', ''))
            
            # Find next empty row
            next_row = ws.max_row + 1
            
            # Write data to Excel
            ws.cell(row=next_row, column=1, value=session_id)  # A: Session ID
            ws.cell(row=next_row, column=2, value=user_prompt)  # B: User Prompt
            ws.cell(row=next_row, column=3, value=system_prompt_full)  # C: System Prompt (Full Text)
            ws.cell(row=next_row, column=4, value=files_str)  # D: Files Accessed
            ws.cell(row=next_row, column=5, value=num_iterations)  # E: Number of Iterations
            ws.cell(row=next_row, column=6, value=script_content_clean)  # F: Script Content (Full Code)
            ws.cell(row=next_row, column=7, value=final_output)  # G: AI Final Output
            # H-J left empty for manual entry
            
            migrated_count += 1
            existing_sessions.add(session_id)  # Add to prevent duplicates in this run
            
        except Exception as e:
            print(f"Error processing {json_file.name}: {e}")
            error_count += 1
            continue
    
    # Save the Excel file
    try:
        wb.save(excel_path)
        print(f"\nMigration completed!")
        print(f"SUCCESS: Migrated {migrated_count} sessions")
        print(f"SKIPPED: {skipped_count} sessions (duplicates)") 
        print(f"ERRORS: {error_count} sessions")
        print(f"\nExcel file updated: {excel_path}")
        
        if migrated_count > 0:
            print(f"\nNext steps:")
            print(f"- Review the migrated data in Excel")
            print(f"- Fill in Expected Output, Human Evaluation, AI Evaluation, and Total Score columns manually")
            print(f"- Future sessions will be automatically added to this file")
    
    except Exception as e:
        print(f"Error saving Excel file: {e}")

def show_migration_summary():
    """Show summary before migration"""
    sessions_dir = Path("query_sessions")
    excel_path = Path("query_tracking.xlsx")
    
    print("=== Session Migration Summary ===")
    
    if sessions_dir.exists():
        json_files = list(sessions_dir.glob("session_*.json"))
        print(f"JSON sessions found: {len(json_files)}")
    else:
        print("No sessions directory found")
        return
    
    if excel_path.exists():
        try:
            wb = load_workbook(excel_path)
            ws = wb.active
            existing_rows = ws.max_row - 1  # Subtract header row
            print(f"Existing Excel entries: {existing_rows}")
        except:
            print("Could not read Excel file")
    else:
        print("Excel tracking file not found")
        return
    
    print("\nProceed with migration? (y/n): ", end="")

if __name__ == "__main__":
    print("Session Data Migration Tool")
    print("===========================")
    
    show_migration_summary()
    
    # Auto-proceed for script execution
    response = input().lower().strip()
    if response in ['y', 'yes', '']:
        migrate_sessions_to_excel()
    else:
        print("Migration cancelled.")
