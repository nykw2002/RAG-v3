#!/usr/bin/env python3
"""
AI File Query System - Unified Backend API
FastAPI backend that integrates all modules and serves the Next.js frontend
"""

import os
import sys
import json
import asyncio
import shutil
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from contextlib import asynccontextmanager

# FastAPI and related imports
from fastapi import FastAPI, HTTPException, UploadFile, File, WebSocket, WebSocketDisconnect, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
import uvicorn

# Our existing modules
from interactive import InteractiveAIFileQuerySystem
from pdf_converter import convert_pdf_to_txt, convert_pdf_to_xml
from migrate_sessions import migrate_sessions_to_excel

# Optional dependencies
try:
    import pandas as pd
except ImportError:
    pd = None
    
try:
    import openpyxl
except ImportError:
    openpyxl = None

# Optional excel integration (restored)
try:
    from excel_integration import append_session_from_json
except ImportError:
    append_session_from_json = None
    print("Warning: Excel integration not available")

# Environment setup
from dotenv import load_dotenv
load_dotenv()

# ============================================================================
# Data Models
# ============================================================================

class ChatMessage(BaseModel):
    id: str
    role: str  # "user" or "assistant"
    content: str
    timestamp: str

class ChatRequest(BaseModel):
    message: str
    session_id: Optional[str] = None

class ChatResponse(BaseModel):
    message: ChatMessage
    session_id: str
    is_complete: bool = False

class SessionInfo(BaseModel):
    id: str
    name: str
    date: str
    messageCount: int
    size: str
    filesAccessed: List[str] = []

class FileInfo(BaseModel):
    id: str
    name: str
    originalName: str
    format: str
    size: str
    date: str
    path: str

class SystemPromptUpdate(BaseModel):
    prompt: str

class ConversionRequest(BaseModel):
    format: str  # "txt" or "xml"

class MigrationStatus(BaseModel):
    isRunning: bool
    progress: int
    logs: List[Dict[str, Any]]
    completed: bool

# ============================================================================
# Global State Management
# ============================================================================

class BackendState:
    def __init__(self):
        self.ai_system = None
        self.active_queries: Dict[str, Any] = {}
        self.migration_status = MigrationStatus(isRunning=False, progress=0, logs=[], completed=False)
        self.websocket_connections: List[WebSocket] = []
    
    async def initialize(self):
        """Initialize AI system and other components"""
        try:
            self.ai_system = InteractiveAIFileQuerySystem()
            print("AI System initialized")
        except Exception as e:
            print(f"Failed to initialize AI system: {e}")
            raise

# Global state instance
state = BackendState()

# Chat persistence utilities
def save_chat_session(session_id: str, messages: List[Dict], user_query: str = None):
    """Save a chat session in frontend-friendly format"""
    try:
        chat_file = Path(f"frontend_chats/chat_{session_id}.json")
        chat_file.parent.mkdir(exist_ok=True)
        
        chat_data = {
            "id": session_id,
            "name": f"Chat {session_id.split('_')[-1]}",
            "messages": messages,
            "created": datetime.now().isoformat(),
            "lastUpdated": datetime.now().isoformat(),
            "user_query": user_query
        }
        
        with open(chat_file, 'w', encoding='utf-8') as f:
            json.dump(chat_data, f, indent=2, ensure_ascii=False)
        
        return chat_file
    except Exception as e:
        print(f"Error saving chat session: {e}")
        return None

def load_chat_sessions():
    """Load all chat sessions from frontend_chats directory"""
    try:
        chats_dir = Path("frontend_chats")
        if not chats_dir.exists():
            return []
        
        sessions = []
        for chat_file in chats_dir.glob("chat_*.json"):
            try:
                with open(chat_file, 'r', encoding='utf-8') as f:
                    chat_data = json.load(f)
                sessions.append(chat_data)
            except Exception as e:
                print(f"Error loading chat {chat_file}: {e}")
                continue
        
        # Sort by creation date (newest first)
        sessions.sort(key=lambda x: x.get('created', ''), reverse=True)
        return sessions
    except Exception as e:
        print(f"Error loading chat sessions: {e}")
        return []

# ============================================================================
# FastAPI App Setup
# ============================================================================

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    await state.initialize()
    yield
    # Shutdown
    print("Shutting down backend...")

app = FastAPI(
    title="AI File Query Backend",
    description="Backend API for AI File Query System",
    version="1.0.0",
    lifespan=lifespan
)

# CORS middleware for Next.js frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:3001"],  # Next.js dev server
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ============================================================================
# WebSocket for Real-time Communication
# ============================================================================

@app.websocket("/ws")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    state.websocket_connections.append(websocket)
    
    try:
        while True:
            # Keep connection alive and handle messages
            data = await websocket.receive_text()
            message = json.loads(data)
            
            if message.get("type") == "ping":
                await websocket.send_text(json.dumps({"type": "pong"}))
    
    except WebSocketDisconnect:
        state.websocket_connections.remove(websocket)
    except Exception as e:
        print(f"WebSocket error: {e}")
        if websocket in state.websocket_connections:
            state.websocket_connections.remove(websocket)

async def broadcast_message(message: Dict[str, Any]):
    """Broadcast message to all connected clients"""
    if state.websocket_connections:
        dead_connections = []
        for connection in state.websocket_connections:
            try:
                await connection.send_text(json.dumps(message))
            except:
                dead_connections.append(connection)
        
        # Remove dead connections
        for conn in dead_connections:
            state.websocket_connections.remove(conn)

# ============================================================================
# Chat API Endpoints
# ============================================================================

@app.post("/api/chat/send", response_model=ChatResponse)
async def send_chat_message(request: ChatRequest, background_tasks: BackgroundTasks):
    """Send a message to the AI and get response"""
    try:
        # Generate session ID if not provided
        session_id = request.session_id or f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{len(state.active_queries):03d}"
        
        # Create user message
        user_message = ChatMessage(
            id=f"msg_{datetime.now().timestamp()}",
            role="user",
            content=request.message,
            timestamp=datetime.now().isoformat()
        )
        
        # Broadcast user message immediately
        await broadcast_message({
            "type": "chat_message",
            "session_id": session_id,
            "message": user_message.dict()
        })
        
        # Process query in background
        background_tasks.add_task(process_chat_query, session_id, request.message)
        
        return ChatResponse(
            message=user_message,
            session_id=session_id,
            is_complete=False
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error processing chat message: {str(e)}")

async def process_chat_query(session_id: str, user_query: str):
    """Process the chat query using our AI system"""
    try:
        # Broadcast that AI is thinking
        await broadcast_message({
            "type": "chat_status",
            "session_id": session_id,
            "status": "thinking"
        })
        
        # Run the query (this is synchronous, so we run it in a thread)
        import asyncio
        loop = asyncio.get_event_loop()
        
        def run_query():
            return state.ai_system.query_files(user_query, session_id)
        
        result, iterations, files_accessed = await loop.run_in_executor(None, run_query)
        
        # Save session with files accessed info
        session_file = state.ai_system.save_session(user_query, iterations, result, session_id, files_accessed)
        
        # Also save in frontend format
        messages = [
            {
                "id": f"msg_{session_id}_user",
                "role": "user",
                "content": user_query,
                "timestamp": datetime.now().isoformat()
            },
            {
                "id": f"msg_{session_id}_assistant", 
                "role": "assistant",
                "content": result,
                "timestamp": datetime.now().isoformat()
            }
        ]
        save_chat_session(session_id, messages, user_query)
        
        # Create AI response message
        ai_message = ChatMessage(
            id=f"msg_{datetime.now().timestamp()}",
            role="assistant",
            content=result,
            timestamp=datetime.now().isoformat()
        )
        
        # Broadcast AI response
        await broadcast_message({
            "type": "chat_message",
            "session_id": session_id,
            "message": ai_message.dict()
        })
        
        # Broadcast completion
        await broadcast_message({
            "type": "chat_complete",
            "session_id": session_id,
            "iterations": len(iterations),
            "files_accessed": files_accessed
        })
        
    except Exception as e:
        # Broadcast error
        await broadcast_message({
            "type": "chat_error",
            "session_id": session_id,
            "error": str(e)
        })

@app.get("/api/chat/sessions/load")
async def load_chat_sessions_api():
    """Load all saved chat sessions for the frontend"""
    try:
        sessions = load_chat_sessions()
        return {"sessions": sessions}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error loading chat sessions: {str(e)}")

@app.post("/api/chat/sessions/save")
async def save_chat_session_api(request: dict):
    """Save a chat session from the frontend"""
    try:
        session_id = request.get("session_id")
        messages = request.get("messages", [])
        user_query = request.get("user_query")
        
        if not session_id:
            raise HTTPException(status_code=400, detail="session_id is required")
        
        chat_file = save_chat_session(session_id, messages, user_query)
        
        if chat_file:
            return {"message": "Chat session saved successfully", "file": str(chat_file)}
        else:
            raise HTTPException(status_code=500, detail="Failed to save chat session")
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving chat session: {str(e)}")

@app.get("/api/chat/sessions", response_model=List[SessionInfo])
async def get_chat_sessions():
    try:
        sessions_dir = Path("query_sessions")
        if not sessions_dir.exists():
            return []
        
        sessions = []
        for session_file in sessions_dir.glob("session_*.json"):
            try:
                with open(session_file, 'r', encoding='utf-8') as f:
                    session_data = json.load(f)
                
                sessions.append(SessionInfo(
                    id=session_data.get('session_id', session_file.stem),
                    name=f"{session_file.name}",
                    date=session_data.get('timestamp', ''),
                    messageCount=session_data.get('total_iterations', 0),
                    size=f"{session_file.stat().st_size / 1024:.1f} KB",
                    filesAccessed=session_data.get('files_accessed', [])
                ))
            except Exception as e:
                print(f"Error reading session {session_file}: {e}")
                continue
        
        # Sort by date (newest first)
        sessions.sort(key=lambda x: x.date, reverse=True)
        return sessions
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching sessions: {str(e)}")

@app.get("/api/chat/sessions/{session_id}")
async def get_session_details(session_id: str):
    """Get detailed session data"""
    try:
        session_file = Path(f"query_sessions/session_{session_id}.json")
        if not session_file.exists():
            raise HTTPException(status_code=404, detail="Session not found")
        
        with open(session_file, 'r', encoding='utf-8') as f:
            session_data = json.load(f)
        
        return session_data
        
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error fetching session: {str(e)}")

@app.delete("/api/chat/sessions/{session_id}")
async def delete_session(session_id: str):
    """Delete a chat session"""
    try:
        session_file = Path(f"query_sessions/session_{session_id}.json")
        if not session_file.exists():
            raise HTTPException(status_code=404, detail="Session not found")
        
        session_file.unlink()
        return {"message": "Session deleted successfully"}
        
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Session not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting session: {str(e)}")

# ============================================================================
# Files API Endpoints
# ============================================================================

@app.get("/api/files/list", response_model=List[FileInfo])
async def list_files():
    """Get list of all processed files"""
    try:
        files_dir = Path("files_to_query")
        if not files_dir.exists():
            return []
        
        files = []
        for file_path in files_dir.iterdir():
            if file_path.is_file() and file_path.suffix.lower() in ['.txt', '.xml', '.pdf']:
                files.append(FileInfo(
                    id=file_path.name,
                    name=file_path.name,
                    originalName=file_path.name,
                    format=file_path.suffix[1:].lower(),
                    size=f"{file_path.stat().st_size / 1024:.1f} KB",
                    date=datetime.fromtimestamp(file_path.stat().st_mtime).isoformat(),
                    path=str(file_path)
                ))
        
        return files
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error listing files: {str(e)}")

@app.post("/api/files/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload a file to the system"""
    try:
        # Validate file type
        allowed_types = {'.pdf', '.txt', '.xml'}
        file_suffix = Path(file.filename).suffix.lower()
        
        if file_suffix not in allowed_types:
            raise HTTPException(status_code=400, detail=f"File type {file_suffix} not allowed")
        
        # Create files directory if it doesn't exist
        files_dir = Path("files_to_query")
        files_dir.mkdir(exist_ok=True)
        
        # Save file
        file_path = files_dir / file.filename
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        return {
            "message": "File uploaded successfully",
            "filename": file.filename,
            "size": file_path.stat().st_size
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error uploading file: {str(e)}")

@app.post("/api/files/{filename}/convert")
async def convert_file(filename: str, request: ConversionRequest):
    """Convert a PDF file to TXT or XML"""
    try:
        file_path = Path(f"files_to_query/{filename}")
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
        
        if file_path.suffix.lower() != '.pdf':
            raise HTTPException(status_code=400, detail="Only PDF files can be converted")
        
        # Determine output path
        output_suffix = f".{request.format}"
        output_path = file_path.parent / f"{file_path.stem}{output_suffix}"
        
        # Convert file
        if request.format == "txt":
            result = convert_pdf_to_txt(str(file_path), str(output_path))
        elif request.format == "xml":
            result = convert_pdf_to_xml(str(file_path), str(output_path))
        else:
            raise HTTPException(status_code=400, detail="Invalid format. Use 'txt' or 'xml'")
        
        if result:
            return {
                "message": f"File converted to {request.format.upper()} successfully",
                "output_file": output_path.name
            }
        else:
            raise HTTPException(status_code=500, detail="Conversion failed")
            
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="File not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error converting file: {str(e)}")

@app.get("/api/files/{filename}/download")
async def download_file(filename: str):
    """Download a file"""
    try:
        file_path = Path(f"files_to_query/{filename}")
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
        
        return FileResponse(
            path=str(file_path),
            filename=filename,
            media_type='application/octet-stream'
        )
        
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="File not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error downloading file: {str(e)}")

@app.delete("/api/files/{filename}")
async def delete_file(filename: str):
    """Delete a file"""
    try:
        file_path = Path(f"files_to_query/{filename}")
        if not file_path.exists():
            raise HTTPException(status_code=404, detail="File not found")
        
        file_path.unlink()
        return {"message": "File deleted successfully"}
        
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="File not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error deleting file: {str(e)}")

# ============================================================================
# System Prompt API Endpoints
# ============================================================================

@app.get("/api/system-prompt")
async def get_system_prompt():
    """Get current system prompt"""
    try:
        prompt_file = Path("system_prompt.txt")
        if prompt_file.exists():
            with open(prompt_file, 'r', encoding='utf-8') as f:
                prompt = f.read()
        else:
            prompt = "Default system prompt not found"
        
        return {"prompt": prompt}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading system prompt: {str(e)}")

@app.put("/api/system-prompt")
async def update_system_prompt(request: SystemPromptUpdate):
    """Update system prompt"""
    try:
        prompt_file = Path("system_prompt.txt")
        with open(prompt_file, 'w', encoding='utf-8') as f:
            f.write(request.prompt)
        
        # Reload AI system with new prompt
        if state.ai_system:
            state.ai_system.system_prompt = request.prompt
        
        return {"message": "System prompt updated successfully"}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating system prompt: {str(e)}")

# ============================================================================
# Evaluation API Endpoints
# ============================================================================

@app.post("/api/evaluation/migrate")
async def start_migration(background_tasks: BackgroundTasks):
    """Start data migration to Excel"""
    if state.migration_status.isRunning:
        raise HTTPException(status_code=400, detail="Migration already in progress")
    
    background_tasks.add_task(run_migration)
    return {"message": "Migration started"}

async def run_migration():
    """Run the migration process with progress tracking"""
    try:
        state.migration_status.isRunning = True
        state.migration_status.progress = 0
        state.migration_status.logs = []
        state.migration_status.completed = False
        
        # Simulate migration steps
        steps = [
            "Initializing migration...",
            "Reading session files...",
            "Processing data...", 
            "Creating Excel file...",
            "Migration completed!"
        ]
        
        for i, step in enumerate(steps):
            # Add log entry
            log_entry = {
                "timestamp": datetime.now().isoformat(),
                "message": step,
                "type": "info"
            }
            state.migration_status.logs.append(log_entry)
            
            # Update progress
            state.migration_status.progress = int((i + 1) / len(steps) * 100)
            
            # Broadcast progress
            await broadcast_message({
                "type": "migration_progress",
                "progress": state.migration_status.progress,
                "logs": state.migration_status.logs
            })
            
            # Simulate work
            await asyncio.sleep(1)
        
        # Actually run migration
        try:
            migrate_sessions_to_excel()
            state.migration_status.completed = True
            
            final_log = {
                "timestamp": datetime.now().isoformat(),
                "message": "Migration completed successfully!",
                "type": "success"
            }
            state.migration_status.logs.append(final_log)
            
        except Exception as e:
            error_log = {
                "timestamp": datetime.now().isoformat(),
                "message": f"Migration failed: {str(e)}",
                "type": "error"
            }
            state.migration_status.logs.append(error_log)
        
        state.migration_status.isRunning = False
        
        # Broadcast final status
        await broadcast_message({
            "type": "migration_complete",
            "completed": state.migration_status.completed,
            "logs": state.migration_status.logs
        })
        
    except Exception as e:
        state.migration_status.isRunning = False
        print(f"Migration error: {e}")

@app.get("/api/evaluation/status", response_model=MigrationStatus)
async def get_migration_status():
    """Get current migration status"""
    return state.migration_status

@app.get("/api/evaluation/download")
async def download_excel():
    """Download the Excel evaluation file"""
    try:
        excel_path = Path("query_tracking.xlsx")
        if not excel_path.exists():
            raise HTTPException(status_code=404, detail="Excel file not found. Run migration first.")
        
        return FileResponse(
            path=str(excel_path),
            filename="evaluation_report.xlsx",
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Excel file not found")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error downloading Excel file: {str(e)}")

@app.get("/api/evaluation/preview")
async def preview_excel():
    """Preview the Excel evaluation file content"""
    try:
        excel_path = Path("query_tracking.xlsx")
        if not excel_path.exists():
            raise HTTPException(status_code=404, detail="Excel file not found. Run migration first.")
        
        # Try to read Excel file using pandas
        try:
            import pandas as pd
            
            # Read the Excel file
            df = pd.read_excel(str(excel_path))
            
            # Handle NaN values - replace with None for JSON serialization
            df = df.fillna(value=None)
            
            # Convert to records (list of dictionaries)
            records = df.to_dict('records')
            
            # Limit to first 50 rows for preview
            preview_records = records[:50]
            
            # Get column names
            columns = df.columns.tolist()
            
            return {
                "columns": columns,
                "data": preview_records,
                "total_rows": len(records),
                "preview_rows": len(preview_records),
                "file_exists": True
            }
            
        except ImportError:
            # Fallback: try to read with openpyxl directly
            try:
                import openpyxl
                workbook = openpyxl.load_workbook(str(excel_path))
                sheet = workbook.active
                
                # Get column names from first row
                columns = [cell.value for cell in sheet[1]]
                
                # Get data rows (limit to 50)
                data = []
                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 1):
                    if row_num > 50:
                        break
                    row_dict = {}
                    for i, value in enumerate(row):
                        if i < len(columns):
                            # Handle various Excel data types that might not be JSON serializable
                            if value is None:
                                row_dict[columns[i]] = None
                            elif isinstance(value, (int, float, str, bool)):
                                # Check for NaN or infinite values
                                if isinstance(value, float) and (value != value or value == float('inf') or value == float('-inf')):
                                    row_dict[columns[i]] = None
                                else:
                                    row_dict[columns[i]] = value
                            else:
                                # Convert other types to string
                                row_dict[columns[i]] = str(value)
                    data.append(row_dict)
                
                return {
                    "columns": columns,
                    "data": data,
                    "total_rows": sheet.max_row - 1,  # Subtract header row
                    "preview_rows": len(data),
                    "file_exists": True
                }
                
            except ImportError:
                raise HTTPException(status_code=500, detail="Neither pandas nor openpyxl available for Excel reading")
        
    except FileNotFoundError:
        return {
            "columns": [],
            "data": [],
            "total_rows": 0,
            "preview_rows": 0,
            "file_exists": False,
            "error": "Excel file not found. Run migration first."
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading Excel file: {str(e)}")

# ============================================================================
# Health Check
# ============================================================================

@app.get("/api/health")
async def health_check():
    """Health check endpoint"""
    return {
        "status": "healthy",
        "timestamp": datetime.now().isoformat(),
        "ai_system": "ready" if state.ai_system else "not_ready"
    }

# ============================================================================
# Main Entry Point
# ============================================================================

if __name__ == "__main__":
    print("Starting AI File Query Backend...")
    print("Ensuring directories exist...")
    
    # Create necessary directories
    Path("files_to_query").mkdir(exist_ok=True)
    Path("query_sessions").mkdir(exist_ok=True)
    Path("temp_scripts").mkdir(exist_ok=True)
    
    print("Backend ready!")
    print("Frontend should connect to: http://localhost:8000")
    print("API docs available at: http://localhost:8000/docs")
    
    uvicorn.run(
        "backend_main:app",
        host="0.0.0.0",
        port=8000,
        reload=False,  # Disabled to prevent interruption during AI processing
        reload_dirs=["./"],
        reload_excludes=["temp_scripts/*", "query_sessions/*", "files_to_query/*", "__pycache__/*", "frontend_chats/*"],
        log_level="info"
    )
