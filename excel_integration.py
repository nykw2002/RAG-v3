#!/usr/bin/env python3
"""
Auto-append new session data to Excel tracking file
Used by interactive system to automatically log each session
"""

import json
import os
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime
import hashlib

def get_system_prompt_version(system_prompt):
    """Get a human-readable version identifier for the system prompt"""
    if not system_prompt:
        return "N/A"
    
    prompt_lower = system_prompt.lower()
    
    # Identify different prompt versions based on key phrases
    if "two-phase process" in prompt_lower:
        return "v2.0 (Two-Phase)"
    elif "critical rule" in prompt_lower:
        return "v1.5 (Critical Rule)"
    elif "script results are authoritative" in prompt_lower:
        return "v1.2 (Authoritative)"
    elif "query_complete" in prompt_lower:
        return "v1.1 (Query Complete)"
    elif "specialized in analyzing" in prompt_lower:
        return "v1.0 (Basic)"
    else:
        # Fallback to hash for unknown versions
        hash_val = hashlib.md5(system_prompt.encode()).hexdigest()[:8]
        return f"Custom ({hash_val})"

def extract_script_content(iterations):
    """Extract the actual script content from iterations"""
    all_scripts = []
    
    for i, iteration in enumerate(iterations, 1):
        if iteration.get('script_executed', False):
            exec_result = iteration.get('execution_result', {})
            script_content = exec_result.get('script_content', '')
            if script_content:
                script_header = f"\n=== Iteration {i} Script ==="
                all_scripts.append(script_header)
                all_scripts.append(script_content)
    
    if all_scripts:
        return '\n'.join(all_scripts)
    else:
        return "No scripts executed"

def clean_text_for_excel(text):
    """Clean text to be Excel-friendly"""
    if not text:
        return ""
    
    # Convert to string and limit length for Excel readability
    text = str(text).strip()
    
    # Replace problematic characters
    text = text.replace('\n', ' | ').replace('\r', '')
    
    # Limit length to avoid Excel issues
    if len(text) > 1000:
        text = text[:997] + "..."
    
    return text

def append_session_to_excel(session_data, excel_path="query_tracking.xlsx"):
    """
    Append a single session to the Excel tracking file
    Returns True if successful, False otherwise
    """
    try:
        # Check if Excel file exists
        if not Path(excel_path).exists():
            print(f"Warning: Excel tracking file {excel_path} not found")
            return False
        
        # Load Excel file
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Check if session already exists
        session_id = session_data.get('session_id', '')
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == session_id:
                print(f"Session {session_id} already exists in Excel, skipping")
                return True
        
        # Extract and clean data
        user_prompt = clean_text_for_excel(session_data.get('user_query', ''))
        system_prompt = session_data.get('system_prompt', '')
        system_prompt_full = clean_text_for_excel(system_prompt)  # Full text
        
        # Files accessed
        files_accessed = session_data.get('files_accessed', [])
        if isinstance(files_accessed, list):
            files_str = ", ".join(files_accessed) if files_accessed else "None detected"
        else:
            files_str = str(files_accessed)
        files_str = clean_text_for_excel(files_str)
        
        # Number of iterations
        num_iterations = session_data.get('total_iterations', 0)
        
        # Script content - actual code
        iterations = session_data.get('iterations', [])
        script_content = extract_script_content(iterations)
        script_content_clean = clean_text_for_excel(script_content)
        
        # AI final output
        final_output = clean_text_for_excel(session_data.get('final_answer', ''))
        
        # Find next empty row
        next_row = ws.max_row + 1
        
        # Write data to Excel
        ws.cell(row=next_row, column=1, value=session_id)  # A: Session ID
        ws.cell(row=next_row, column=2, value=user_prompt)  # B: User Prompt
        ws.cell(row=next_row, column=3, value=system_prompt_full)  # C: System Prompt (Full Text)
        ws.cell(row=next_row, column=4, value=files_str)  # D: Files Accessed
        ws.cell(row=next_row, column=5, value=num_iterations)  # E: Number of Iterations
        ws.cell(row=next_row, column=6, value=script_content_clean)  # F: Script Content (Full Code)
        ws.cell(row=next_row, column=7, value=final_output)  # G: AI Final Output
        # H-J left empty for manual entry
        
        # Save the file
        wb.save(excel_path)
        print(f"SUCCESS: Added session {session_id} to Excel tracking")
        return True
        
    except Exception as e:
        print(f"Error appending session to Excel: {e}")
        return False

def append_session_from_json(json_file_path, excel_path="query_tracking.xlsx"):
    """
    Load session data from JSON file and append to Excel
    """
    try:
        with open(json_file_path, 'r', encoding='utf-8') as f:
            session_data = json.load(f)
        return append_session_to_excel(session_data, excel_path)
    except Exception as e:
        print(f"Error loading session from {json_file_path}: {e}")
        return False

if __name__ == "__main__":
    # Test function - append the latest session
    sessions_dir = Path("query_sessions")
    if sessions_dir.exists():
        json_files = sorted(sessions_dir.glob("session_*.json"), key=lambda x: x.stat().st_mtime)
        if json_files:
            latest_session = json_files[-1]
            print(f"Testing with latest session: {latest_session.name}")
            success = append_session_from_json(latest_session)
            if success:
                print("Test successful!")
            else:
                print("Test failed!")
        else:
            print("No session files found")
    else:
        print("No sessions directory found")
